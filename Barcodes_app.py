
import io
import pandas as pd
import streamlit as st
import barcode
from barcode.writer import ImageWriter
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment

st.set_page_config(page_title="Barcode Generator", page_icon="📦", layout="wide")

st.markdown("""
<style>
.hero{padding:30px;border-radius:12px;background:linear-gradient(90deg,#0f172a,#2563eb);color:white}
.small{color:#666}
</style>
<div class="hero">
<h1>📦 Barcode Generator</h1>
<p>Faça upload de qualquer arquivo Excel, escolha a coluna correta e baixe um novo Excel com os códigos de barras.</p>
</div>
""", unsafe_allow_html=True)

st.markdown("""
### Como usar
1. Envie um arquivo Excel (.xlsx)
2. Confira a prévia.
3. Escolha a coluna que contém os Waybills (o sistema tenta identificar automaticamente).
4. Clique em **Gerar**.
5. Baixe o novo Excel.
""")

uploaded = st.file_uploader("Selecione um arquivo Excel", type=["xlsx"])

def guess(cols):
    names=["waybills","waybill","tracking","tracking number","awb","awb number","shipment","shipment number"]
    lower={c.lower():c for c in cols}
    for n in names:
        if n in lower:
            return lower[n]
    return cols[0]

if uploaded:
    try:
        df=pd.read_excel(uploaded)
    except Exception as e:
        st.error(f"Não foi possível abrir o Excel: {e}")
        st.stop()

    st.success(f"Arquivo carregado com {len(df)} linhas.")
    st.dataframe(df.head())

    col=st.selectbox("Coluna com os Waybills", df.columns,
                     index=list(df.columns).index(guess(list(df.columns))))

    if st.button("🚀 Gerar códigos de barras"):
        wb=Workbook()
        ws=wb.active
        ws.title="Barcode Report"
        ws["A1"]="Waybill"
        ws["B1"]="Barcode"
        f=Font(bold=True,size=12)
        ws["A1"].font=f
        ws["B1"].font=f
        ws.column_dimensions["A"].width=28
        ws.column_dimensions["B"].width=34

        Code128=barcode.get_barcode_class("code128")
        prog=st.progress(0,text="Gerando...")

        total=max(len(df),1)
        for i,val in enumerate(df[col].fillna("").astype(str),start=2):
            ws[f"A{i}"]=val
            ws[f"A{i}"].alignment=Alignment(vertical="center")

            buf=io.BytesIO()
            Code128(val,writer=ImageWriter()).write(
                buf,
                options={"module_width":0.2,"module_height":7,
                         "font_size":10,"text_distance":5}
            )
            buf.seek(0)
            img=OpenpyxlImage(buf)
            img.width=200
            img.height=90
            img.anchor=f"B{i}"
            ws.add_image(img)
            ws.row_dimensions[i].height=80

            prog.progress((i-1)/total,text=f"Gerando {i-1}/{total}")

        out=io.BytesIO()
        wb.save(out)
        out.seek(0)

        st.success("Arquivo gerado com sucesso!")
        st.download_button(
            "📥 Baixar Excel",
            data=out,
            file_name="AWB_Barcodes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

st.markdown("---")
st.caption("Aplicação exemplo em Streamlit usando pandas + python-barcode + openpyxl.")
